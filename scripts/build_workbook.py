#!/usr/bin/env python3
"""Build plan/2026-2027/APEX-Yearly-Plan.xlsx from the roster and the rotation.

Six sheets: Overview, Call Schedule, Roster, KPI Tracker, Q1 90-Day Test,
Projects. The counts on Overview and the KPI Tracker are live formulas over
the other sheets, so marking a call done or filling a roster row moves the
numbers — the workbook is meant to be worked in, not read once.

Regenerating overwrites the file. Anything typed into it is lost, so it takes
--force once the file exists.

    python3 scripts/build_workbook.py [--force]
"""

import argparse
import csv
import datetime as dt
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
ROSTER = ROOT / "config" / "network" / "roster.csv"
SCHEDULE = ROOT / "plan" / "2026-2027" / "call-schedule.csv"
CONFIG = ROOT / "config" / "network" / "network.yaml"
TARGET = ROOT / "plan" / "2026-2027" / "APEX-Yearly-Plan.xlsx"

FONT = "Arial"
INK = "1F2937"
BRASS = "96661C"
MUTED = "6B7280"

H1 = Font(name=FONT, size=16, bold=True, color=INK)
H2 = Font(name=FONT, size=12, bold=True, color=INK)
HEAD = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY = Font(name=FONT, size=10, color=INK)
SMALL = Font(name=FONT, size=9, color=MUTED)
ACCENT = Font(name=FONT, size=10, bold=True, color=BRASS)
BIG = Font(name=FONT, size=20, bold=True, color=BRASS)

HEAD_FILL = PatternFill("solid", fgColor="374151")
BAND = PatternFill("solid", fgColor="F3F4F6")
# Yellow marks every cell a person is expected to type into.
INPUT = PatternFill("solid", fgColor="FFF9C4")
TIER_FILL = {"A": PatternFill("solid", fgColor="F5E6C8"),
             "B": PatternFill("solid", fgColor="DCE6F0"),
             "C": PatternFill("solid", fgColor="DDE7DE")}

THIN = Side(style="thin", color="D1D5DB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def header_row(ws, row, headers, widths):
    for i, (text, width) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=text)
        c.font, c.fill, c.border = HEAD, HEAD_FILL, BOX
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def load():
    roster = list(csv.DictReader(ROSTER.open(encoding="utf-8")))
    schedule = list(csv.DictReader(SCHEDULE.open(encoding="utf-8")))
    config = yaml.safe_load(CONFIG.read_text(encoding="utf-8"))
    return roster, schedule, config


# ---------------------------------------------------------------- Call Schedule
def sheet_schedule(wb, schedule):
    ws = wb.create_sheet("Call Schedule")
    cols = ["Date", "Day", "Wk", "Qtr", "Who", "Cohort", "Tier", "Channel",
            "What the call is for", "Phone", "Email", "Status", "Notes from the call"]
    header_row(ws, 1, cols, [11, 10, 5, 6, 22, 8, 6, 10, 62, 16, 24, 11, 46])

    status = DataValidation(type="list",
                            formula1='"done,rescheduled,no answer,skipped"',
                            allow_blank=True)
    ws.add_data_validation(status)

    for r, row in enumerate(schedule, start=2):
        ws.cell(row=r, column=1, value=dt.date.fromisoformat(row["date"])
                ).number_format = "dd/mm/yyyy"
        ws.cell(row=r, column=2, value=row["weekday"][:3])
        ws.cell(row=r, column=3, value=int(row["week_no"]))
        ws.cell(row=r, column=4, value=row["quarter"])
        ws.cell(row=r, column=5, value=row["name"])
        ws.cell(row=r, column=6, value=row["cohort"])
        t = ws.cell(row=r, column=7, value=row["tier"])
        t.fill = TIER_FILL[row["tier"]]
        t.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=8, value=row["channel"])
        ws.cell(row=r, column=9, value=row["purpose"]).alignment = WRAP
        ws.cell(row=r, column=10, value=row["phone"])
        ws.cell(row=r, column=11, value=row["email"])
        for col in (12, 13):                      # Status and Notes are typed in
            ws.cell(row=r, column=col).fill = INPUT
        status.add(ws.cell(row=r, column=12))
        for col in range(1, 14):
            c = ws.cell(row=r, column=col)
            c.border = BOX
            if c.font.name != FONT:
                c.font = BODY
            if col != 9:
                c.alignment = TOP
        ws.row_dimensions[r].height = 30

    last = len(schedule) + 1
    ws.auto_filter.ref = f"A1:M{last}"
    ws.conditional_formatting.add(
        f"L2:L{last}",
        CellIsRule(operator="equal", formula=['"done"'],
                   fill=PatternFill("solid", fgColor="D1FAE5")))
    ws.cell(row=1, column=12).comment = Comment(
        "Type the outcome here. 'done' turns the row green and feeds every "
        "count on Overview and the KPI Tracker.", "APEX plan")
    return last


# ---------------------------------------------------------------------- Roster
ROSTER_COLS = [
    ("id", "ID", 5, False), ("name", "Name", 22, False),
    ("name_he", "שם בעברית", 16, True), ("cohort", "Cohort", 8, False),
    ("founder_track", "Founder track", 12, False), ("track", "Unit", 12, False),
    ("company", "Company", 22, False), ("role", "Role", 30, False),
    ("location", "Location", 14, True), ("domain", "Domain", 16, False),
    ("tier", "Tier", 6, False), ("tie_strength", "Tie strength", 13, True),
    ("current_goal", "Now / current goal", 34, True),
    ("needs", "Needs", 30, True), ("can_help", "Can help with", 30, True),
    ("last_contact", "Last contact", 12, True),
    ("last_source", "Source", 16, True),
    ("next_action", "NEXT ACTION", 34, True),
    ("next_action_due", "Due", 11, True),
    ("linkedin", "LinkedIn", 30, False), ("email", "Email", 24, True),
    ("phone", "Phone", 16, True),
    ("source_note", "Source note (GPU screen, 15/07)", 46, False),
    ("notes", "Notes", 40, True),
]


def sheet_roster(wb, roster):
    ws = wb.create_sheet("Roster")
    header_row(ws, 1, [c[1] for c in ROSTER_COLS], [c[2] for c in ROSTER_COLS])

    tie = DataValidation(
        type="list",
        formula1='"Known,Connected,Engaged,Contributor,Node"', allow_blank=True)
    ws.add_data_validation(tie)

    for r, row in enumerate(roster, start=2):
        for i, (key, _, _, editable) in enumerate(ROSTER_COLS, start=1):
            val = row[key]
            if key == "id":
                val = int(val) if val else None
            c = ws.cell(row=r, column=i, value=val or None)
            c.font, c.border, c.alignment = BODY, BOX, TOP
            if editable:
                c.fill = INPUT
            if key == "tier":
                c.fill = TIER_FILL[row["tier"]]
                c.alignment = Alignment(horizontal="center", vertical="top")
            if key in ("role", "source_note", "current_goal", "needs",
                       "can_help", "next_action", "notes"):
                c.alignment = WRAP
            if key == "last_contact" or key == "next_action_due":
                c.number_format = "dd/mm/yyyy"
        tie.add(ws.cell(row=r, column=12))
        ws.row_dimensions[r].height = 30

    last = len(roster) + 1
    ws.auto_filter.ref = f"A1:X{last}"
    ws.cell(row=1, column=18).comment = Comment(
        "The most important field in the workbook. Role definition §7: "
        "'a database without a next action is just a contact list.' "
        "A row with no next action is stored, not maintained.", "APEX plan")
    ws.cell(row=1, column=23).comment = Comment(
        "Verbatim from the 15/07/26 master sheet, which screened everyone "
        "against GPU-systems / power / real-time-control criteria for a "
        "hardware founding team. Kept because it is real research — but it "
        "answers a different question than this role asks, and nothing here "
        "uses it to rank or deprioritise anyone.", "APEX plan")
    return last


# -------------------------------------------------------------------- Overview
def sheet_overview(wb, roster, schedule, config, n_roster, n_sched):
    ws = wb.create_sheet("Overview", 0)
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDEFGH", [26, 13, 13, 13, 13, 13, 13, 30]):
        ws.column_dimensions[col].width = width

    ws["A1"] = "APEX — Talent, Alumni & Network"
    ws["A1"].font = H1
    ws["A2"] = "Yearly plan · September 2026 – August 2027"
    ws["A2"].font = Font(name=FONT, size=11, color=MUTED)
    ws["A3"] = ("Owner: Nadav Delgo   ·   Reviewed weekly with Avishag   ·   "
                "Built from the role definition (§ references point at it)")
    ws["A3"].font = SMALL

    ws["A5"] = ("אתה ה־Human Node של APEX. וככל שאתה מצליח יותר בתפקיד, "
                "אני צריכה להיות פחות באמצע.")
    ws["A5"].font = Font(name=FONT, size=11, italic=True, color=INK)
    ws.merge_cells("A5:H5")
    ws["A5"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[5].height = 22
    ws["A6"] = "— the role definition, §25"
    ws["A6"].font = SMALL

    # --- live headline numbers ------------------------------------------------
    ws["A8"] = "Where things stand"
    ws["A8"].font = H2
    tiles = [
        ("A", "Alumni", f"=COUNTA(Roster!$B$2:$B${n_roster})"),
        ("C", "Conversations planned", f"=COUNTA('Call Schedule'!$E$2:$E${n_sched})"),
        ("E", "Done so far",
         f'=COUNTIF(\'Call Schedule\'!$L$2:$L${n_sched},"done")'),
        ("G", "Roster rows with a next action",
         f"=COUNTA(Roster!$R$2:$R${n_roster})"),
    ]
    for col, label, formula in tiles:
        ws[f"{col}9"] = formula
        ws[f"{col}9"].font = BIG
        ws[f"{col}10"] = label
        ws[f"{col}10"].font = SMALL
        ws[f"{col}10"].alignment = WRAP
    ws.row_dimensions[9].height = 30
    ws.row_dimensions[10].height = 26

    # --- tiers ----------------------------------------------------------------
    ws["A13"] = "Who the alumni are"
    ws["A13"].font = H2
    ws["A14"] = ("Tier sets call frequency and nothing else. It is not a ranking, "
                 "and the matching engine ignores it — a tier-C engineer at Wiz is "
                 "still a first-class referral path into Wiz.")
    ws["A14"].font = SMALL
    ws.merge_cells("A14:H14")
    ws["A14"].alignment = WRAP
    ws.row_dimensions[14].height = 26

    head = ["Tier", "Who", "Calls/yr", "People", "Calls/yr total",
            "Done", "% done"]
    for i, h in enumerate(head, start=1):
        c = ws.cell(row=16, column=i, value=h)
        c.font, c.fill, c.border = HEAD, HEAD_FILL, BOX
    tier_who = {
        "A": "Founders, stealth builders, people at a transition point (§5)",
        "B": "Seniority or research standing — office hours, referrals, intros",
        "C": "No verified role (the roll-up backlog) + slower-moving ICs",
    }
    for j, tier in enumerate("ABC"):
        r = 17 + j
        per_year = config["tiers"][tier]["calls_per_year"]
        ws.cell(row=r, column=1, value=tier).fill = TIER_FILL[tier]
        ws.cell(row=r, column=1).font = ACCENT
        ws.cell(row=r, column=2, value=tier_who[tier]).alignment = WRAP
        ws.cell(row=r, column=3, value=per_year)
        ws.cell(row=r, column=4,
                value=f'=COUNTIF(Roster!$K$2:$K${n_roster},$A{r})')
        ws.cell(row=r, column=5,
                value=f"=COUNTIF('Call Schedule'!$G$2:$G${n_sched},$A{r})")
        ws.cell(row=r, column=6,
                value=f"=COUNTIFS('Call Schedule'!$G$2:$G${n_sched},$A{r},"
                      f"'Call Schedule'!$L$2:$L${n_sched},\"done\")")
        ws.cell(row=r, column=7, value=f"=IF($E{r}=0,0,$F{r}/$E{r})")
        ws.cell(row=r, column=7).number_format = "0%"
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = BOX
            if col != 2:
                ws.cell(row=r, column=col).alignment = Alignment(
                    horizontal="center", vertical="top")
        ws.row_dimensions[r].height = 26

    r = 20
    ws.cell(row=r, column=1, value="Total").font = H2
    for col, formula in ((4, f"=SUM($D$17:$D$19)"), (5, "=SUM($E$17:$E$19)"),
                         (6, "=SUM($F$17:$F$19)")):
        c = ws.cell(row=r, column=col, value=formula)
        c.font = H2
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=7, value="=IF($E20=0,0,$F20/$E20)")
    ws.cell(row=r, column=7).number_format = "0%"
    ws.cell(row=r, column=7).font = H2
    for col in range(1, 8):
        ws.cell(row=r, column=col).border = BOX

    ws["A22"] = ("Rotation: 3–4 conversations a week, Sunday to Wednesday, "
                 "across 50 working weeks. Thursday stays clear for the weekly "
                 "report. חגים are excluded — no call is scheduled on one.")
    ws["A22"].font = SMALL
    ws.merge_cells("A22:H22")
    ws["A22"].alignment = WRAP

    # --- quarters -------------------------------------------------------------
    ws["A25"] = "The year in four movements"
    ws["A25"].font = H2
    qhead = ["Quarter", "Theme", "The test it has to pass", "Calls", "Done",
             "% done"]
    for i, h in enumerate(qhead, start=1):
        c = ws.cell(row=26, column=i, value=h)
        c.font, c.fill, c.border = HEAD, HEAD_FILL, BOX
    quarters = [
        ("Q1 · Sep–Nov 2026", "Know them",
         "The 90-day test (§24) — every question answerable without asking anyone"),
        ("Q2 · Dec–Feb", "Harness them",
         "Alumni move from receiving value to giving it (§4)"),
        ("Q3 · Mar–May 2027", "Compound",
         "Alumni-to-alumni introductions outnumber the ones Nadav brokers"),
        ("Q4 · Jun–Aug 2027", "Institutionalise",
         "A month without Avishag, run for real — does the network still strengthen?"),
    ]
    for j, (label, theme, test) in enumerate(quarters):
        r = 27 + j
        q = f"Q{j + 1}"
        ws.cell(row=r, column=1, value=label).font = ACCENT
        ws.cell(row=r, column=2, value=theme)
        ws.cell(row=r, column=3, value=test).alignment = WRAP
        ws.cell(row=r, column=4,
                value=f"=COUNTIF('Call Schedule'!$D$2:$D${n_sched},\"{q}\")")
        ws.cell(row=r, column=5,
                value=f"=COUNTIFS('Call Schedule'!$D$2:$D${n_sched},\"{q}\","
                      f"'Call Schedule'!$L$2:$L${n_sched},\"done\")")
        ws.cell(row=r, column=6, value=f"=IF($D{r}=0,0,$E{r}/$D{r})")
        ws.cell(row=r, column=6).number_format = "0%"
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BOX
            if col >= 4:
                ws.cell(row=r, column=col).alignment = Alignment(
                    horizontal="center", vertical="top")
        ws.row_dimensions[r].height = 30

    # --- legend ---------------------------------------------------------------
    ws["A33"] = "How to use this workbook"
    ws["A33"].font = H2
    legend = [
        ("Yellow cells are yours to fill in.", "Everything else is generated "
         "from config/network/roster.csv and the call rotation, and is "
         "overwritten when the workbook is rebuilt."),
        ("Call Schedule", "One row per planned conversation. Set Status to "
         "'done' — the row turns green and every count above moves."),
        ("Roster", "The source of truth for people. After each call fill Now, "
         "Needs, Can help, Last contact and above all NEXT ACTION."),
        ("KPI Tracker", "Weekly numbers. The call counts and the funnel "
         "compute themselves; the rest you type."),
        ("Q1 90-Day Test", "The eight questions §24 says must be answerable "
         "in ninety days."),
        ("Projects", "Coffee Club, SF, Office Hours, Alumni Roll-Up — with dates."),
    ]
    r = 34
    for label, text in legend:
        ws.cell(row=r, column=1, value=label).font = ACCENT
        c = ws.cell(row=r, column=2, value=text)
        c.font, c.alignment = SMALL, WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 26
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Example of a filled roster row").font = H2
    r += 1
    example = [
        ("Now / current goal", "Leaving Meta in November, starting a company"),
        ("Needs", "Technical co-founder; two intros in AI infra"),
        ("Can help with", "ML pipelines; interviewing; referrals into Meta"),
        ("Tie strength", "Engaged"),
        ("Last contact", "14/09/2026"),
        ("Source", "call 14/09"),
        ("NEXT ACTION", "Intro to Guy Goldenberg (Wiz) on the ML infra opening"),
        ("Due", "21/09/2026"),
    ]
    for label, value in example:
        ws.cell(row=r, column=1, value=label).font = SMALL
        c = ws.cell(row=r, column=2, value=value)
        c.font, c.fill, c.border = BODY, INPUT, BOX
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1
    ws.cell(row=r, column=1,
            value="Illustration only — this is not a real alum, and no row "
                  "like it exists in the Roster sheet.").font = SMALL

    # --- provenance -----------------------------------------------------------
    r += 3
    ws.cell(row=r, column=1, value="Where the data came from").font = H2
    notes = [
        f"Alumni: {len(roster)} rows from Apex_Master_150726.xlsx "
        f"(cohort #1 26, #2 33, #3 38).",
        "That sheet screened alumni against GPU-systems / power / real-time "
        "criteria for a hardware founding team. Its verdicts are preserved in "
        "the Roster's Source note column and are not used for tiering here.",
        "Tiers, cadence, KPI targets and project definitions: "
        "config/network/network.yaml, derived from the role definition.",
        "חגים read from the 'חגים בישראל' calendar, not inferred. Only "
        "Sep–Nov 2026 is filled in; extend cadence.off_days to cover the rest.",
        "Open with Avishag: 97 alumni here vs 82 graduates in the Canon vs the "
        "117 quoted to her on 26/08 — the funnel needs one denominator.",
    ]
    for text in notes:
        r += 1
        c = ws.cell(row=r, column=1, value="• " + text)
        c.font, c.alignment = SMALL, WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 26
    return ws


# ---------------------------------------------------------------- KPI Tracker
def sheet_kpi(wb, config, n_roster, n_sched, schedule):
    ws = wb.create_sheet("KPI Tracker")
    rep = config["reporting"]["thursday_report"]

    ws["A1"] = "Weekly KPIs"
    ws["A1"].font = H1
    ws["A2"] = ("§17 — not every number rises every week. We are looking for "
                "trend and quality. Grey columns compute themselves from the "
                "Call Schedule and the Roster; yellow ones you fill in on "
                "Thursday.")
    ws["A2"].font = SMALL
    ws.merge_cells("A2:L2")
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 28

    cols = [
        ("Week of", 12), ("Qtr", 6), ("Calls planned", 11), ("Calls done", 11),
        ("Alumni info updated", 13), ("Quality connections", 12),
        ("Connections → outcome", 13), ("Alumni who helped an alum", 14),
        ("Office Hours contributors", 14), ("New SF relationships", 13),
        ("Moved Avishag → Nadav", 14), ("Needed Avishag", 12),
        ("What I learned / recommend", 52),
    ]
    header_row(ws, 4, [c[0] for c in cols], [c[1] for c in cols])

    # Target row, straight from network.yaml so the sheet cannot drift from it.
    k = rep["kpis"]
    targets = [None, None, None, k["meaningful_alumni_calls"]["week"],
               k["alumni_with_current_info"]["week"],
               k["quality_connections_made"]["week"],
               k["connections_that_produced_an_outcome"]["week"],
               k["alumni_who_helped_another_alum"]["week"],
               k["office_hours_contributors"]["week"],
               k["new_sf_relationships"]["week"],
               k["relationships_moved_from_avishag_to_nadav"]["week"],
               k["things_that_genuinely_needed_avishag"]["week"], None]
    ws.cell(row=5, column=1, value="Weekly target").font = ACCENT
    for i, t in enumerate(targets, start=1):
        c = ws.cell(row=5, column=i)
        if i > 2 and t is not None:
            c.value = t
            c.alignment = Alignment(horizontal="center")
        c.font, c.fill, c.border = ACCENT, BAND, BOX
    ws.cell(row=5, column=1).value = "Weekly target"

    weeks = sorted({dt.date.fromisoformat(r["date"]) -
                    dt.timedelta(days=(dt.date.fromisoformat(r["date"]).weekday() + 1) % 7)
                    for r in schedule})
    quarters = {dt.date.fromisoformat(r["date"]): r["quarter"] for r in schedule}

    for j, monday in enumerate(weeks):
        r = 6 + j
        ws.cell(row=r, column=1, value=monday).number_format = "dd/mm/yyyy"
        q = next((quarters[d] for d in quarters
                  if 0 <= (d - monday).days <= 6), "")
        ws.cell(row=r, column=2, value=q).alignment = Alignment(
            horizontal="center")
        ws.cell(row=r, column=3,
                value=f"=COUNTIFS('Call Schedule'!$A$2:$A${n_sched},\">=\"&$A{r},"
                      f"'Call Schedule'!$A$2:$A${n_sched},\"<=\"&$A{r}+6)")
        ws.cell(row=r, column=4,
                value=f"=COUNTIFS('Call Schedule'!$A$2:$A${n_sched},\">=\"&$A{r},"
                      f"'Call Schedule'!$A$2:$A${n_sched},\"<=\"&$A{r}+6,"
                      f"'Call Schedule'!$L$2:$L${n_sched},\"done\")")
        ws.cell(row=r, column=5,
                value=f"=COUNTIFS(Roster!$P$2:$P${n_roster},\">=\"&$A{r},"
                      f"Roster!$P$2:$P${n_roster},\"<=\"&$A{r}+6)")
        for col in (3, 4, 5):
            ws.cell(row=r, column=col).fill = BAND
        for col in range(6, 13):                    # typed in on Thursday
            ws.cell(row=r, column=col).fill = INPUT
        ws.cell(row=r, column=13).fill = INPUT
        for col in range(1, 14):
            c = ws.cell(row=r, column=col)
            c.border, c.font = BOX, BODY
            c.alignment = WRAP if col == 13 else Alignment(
                horizontal="center", vertical="top")
        ws.row_dimensions[r].height = 22

    last = 5 + len(weeks)
    ws.cell(row=4, column=5).comment = Comment(
        "Counts roster rows whose Last contact falls in the week. Fill Last "
        "contact after every conversation and this moves on its own.",
        "APEX plan")

    # --- funnel and dependency ------------------------------------------------
    base = last + 3
    ws.cell(row=base, column=1, value="The funnel (§19)").font = H2
    ws.cell(row=base + 1, column=1,
            value="Counted off Tie strength in the Roster. Far more meaningful "
                  "than 'I did 17 calls'. Early on it will look thin — report "
                  "it thin.").font = SMALL
    ws.merge_cells(start_row=base + 1, start_column=1, end_row=base + 1,
                   end_column=6)
    ws.cell(row=base + 1, column=1).alignment = WRAP

    stages = [("Alumni", f"=COUNTA(Roster!$B$2:$B${n_roster})"),
              ("Known", f'=COUNTIF(Roster!$L$2:$L${n_roster},"Known")'),
              ("Connected", f'=COUNTIF(Roster!$L$2:$L${n_roster},"Connected")'),
              ("Engaged", f'=COUNTIF(Roster!$L$2:$L${n_roster},"Engaged")'),
              ("Contributor", f'=COUNTIF(Roster!$L$2:$L${n_roster},"Contributor")'),
              ("Node", f'=COUNTIF(Roster!$L$2:$L${n_roster},"Node")')]
    for i, (label, formula) in enumerate(stages):
        c1 = ws.cell(row=base + 3, column=1 + i, value=label)
        c1.font, c1.fill, c1.border = HEAD, HEAD_FILL, BOX
        c1.alignment = Alignment(horizontal="center")
        c2 = ws.cell(row=base + 4, column=1 + i, value=formula)
        c2.font, c2.border = BIG, BOX
        c2.alignment = Alignment(horizontal="center")
    ws.row_dimensions[base + 4].height = 30

    d = base + 6
    ws.cell(row=d, column=1, value="Dependency on Avishag (§18)").font = H2
    ws.cell(row=d + 1, column=1,
            value="What share of the week's items closed without her. 100% is "
                  "not the goal — less over time is.").font = SMALL
    ws.merge_cells(start_row=d + 1, start_column=1, end_row=d + 1, end_column=6)
    for i, (label, val) in enumerate([
            ("Items handled this week", None),
            ("Of those, closed without Avishag", None),
            ("% handled independently", f"=IF($B{d + 3}=0,0,$B{d + 4}/$B{d + 3})")]):
        r = d + 3 + i
        ws.cell(row=r, column=1, value=label).font = BODY
        c = ws.cell(row=r, column=2, value=val)
        c.border = BOX
        if val is None:
            c.fill = INPUT
        else:
            c.number_format = "0%"
            c.font = ACCENT

    t = d + 8
    ws.cell(row=t, column=1, value="Three lines, every Thursday (§21)").font = H2
    for i, line in enumerate([
            "מה למדתי — what did I learn about our people this week?",
            "מה אני ממליץ — what should APEX do about it?",
            "מה אני צריך מאבישג — max three. Zero is better."]):
        r = t + 1 + i
        ws.cell(row=r, column=1, value=line).font = BODY
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(row=r, column=4)
        c.fill, c.border = INPUT, BOX
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=13)
        ws.row_dimensions[r].height = 24


# ------------------------------------------------------------- Q1 90-Day Test
def sheet_ninety(wb, n_roster, n_sched):
    ws = wb.create_sheet("Q1 90-Day Test")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "The 90-day test"
    ws["A1"].font = H1
    ws["A2"] = ("§24 — in ninety days these must be answerable without asking "
                "anyone. Q1 is built backwards from them. Deadline: 30 Nov 2026.")
    ws["A2"].font = SMALL
    ws.merge_cells("A2:E2")
    ws["A2"].alignment = WRAP

    header_row(ws, 4, ["#", "The question §24 asks",
                       "What has to exist for it to be answerable",
                       "Where it is tracked", "Status"],
               [5, 44, 46, 26, 14])

    items = [
        ("Who are our 20 strongest alumni?",
         "All 27 tier-A alumni called at least once",
         "Call Schedule, filter Tier = A"),
        ("What is each of them doing right now?",
         "Now / current goal filled for every tier-A row", "Roster, column M"),
        ("Who is on the way to founding?",
         "The founding-track question asked on every tier-A call",
         "Roster, Notes + Now"),
        ("What does each of them need?", "Needs filled for every tier-A row",
         "Roster, column N"),
        ("What can each of them help with?",
         "Can help with filled for every tier-A row", "Roster, column O"),
        ("Who already helps others?",
         "Tie strength set to Contributor or Node where true",
         "Roster, column L"),
        ("Who are our people in SF?",
         "SF phases 1–2 done — circle mapped, key people met",
         "Projects sheet"),
        ("Do we have the first 20 Office Hours contributors?",
         "8 recruited by 30 Nov, with the anti-abuse model researched first",
         "Projects sheet"),
        ("Which relationships moved from Avishag to Nadav?",
         "Counted weekly from week one, so a trend exists by Q4",
         "KPI Tracker, column K"),
        ("The 16 alumni with no verified role",
         "Role and company established for each", "Roster, filter Role = blank"),
        ("If Avishag vanished for a month, would the network keep strengthening?",
         "The honest answer, written down", "Thursday report"),
    ]
    status = DataValidation(type="list",
                            formula1='"not started,in progress,done"',
                            allow_blank=True)
    ws.add_data_validation(status)
    for i, (q, need, where) in enumerate(items, start=1):
        r = 4 + i
        ws.cell(row=r, column=1, value=i).alignment = Alignment(
            horizontal="center", vertical="top")
        ws.cell(row=r, column=2, value=q).alignment = WRAP
        ws.cell(row=r, column=3, value=need).alignment = WRAP
        ws.cell(row=r, column=4, value=where).alignment = WRAP
        c = ws.cell(row=r, column=5)
        c.fill, c.alignment = INPUT, Alignment(horizontal="center")
        status.add(c)
        for col in range(1, 6):
            cc = ws.cell(row=r, column=col)
            cc.border = BOX
            if cc.font.name != FONT:
                cc.font = BODY
        ws.row_dimensions[r].height = 32

    r = 6 + len(items)
    ws.cell(row=r, column=2, value="Tier-A calls completed").font = ACCENT
    ws.cell(row=r, column=3,
            value=f"=COUNTIFS('Call Schedule'!$G$2:$G${n_sched},\"A\","
                  f"'Call Schedule'!$L$2:$L${n_sched},\"done\")"
            ).font = ACCENT
    ws.cell(row=r + 1, column=2,
            value="Roster rows with Needs filled").font = ACCENT
    ws.cell(row=r + 1, column=3,
            value=f"=COUNTA(Roster!$N$2:$N${n_roster})").font = ACCENT
    ws.cell(row=r + 2, column=2,
            value="Roster rows still with no verified role").font = ACCENT
    ws.cell(row=r + 2, column=3,
            value=f'=COUNTIF(Roster!$H$2:$H${n_roster},"Unknown*")').font = ACCENT


# ------------------------------------------------------------------- Projects
def sheet_projects(wb, config):
    ws = wb.create_sheet("Projects")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Standing projects"
    ws["A1"].font = H1
    ws["A2"] = "§9–§12. Dates are the plan's; owners and status are yours."
    ws["A2"].font = SMALL

    header_row(ws, 4, ["Project", "What it is", "Target", "When", "Status",
                       "Notes"], [20, 56, 22, 20, 14, 40])

    status = DataValidation(type="list",
                            formula1='"not started,in progress,done,parked"',
                            allow_blank=True)
    ws.add_data_validation(status)

    rows = [
        ("Alumni Roll-Up",
         "A live picture of every alum — not a list of names. Done when every "
         "row carries role, company, current goal and a next action.",
         "16 unknowns → 0", "By 30 Nov 2026"),
        ("Coffee Club #1", "Four alumni, one per cohort, plus Avishag and Nadav. "
         "Listen through them, and start making each one active in the network. "
         "Nobody repeats within the year. Not called Coffee Club out loud.",
         "4 alumni", "Early Oct 2026"),
        ("Coffee Club #2", "As above — follow-ups from #1 owned by Nadav.",
         "4 alumni", "Early Dec 2026"),
        ("Coffee Club #3", "As above.", "4 alumni", "Feb 2027"),
        ("Coffee Club #4", "As above.", "4 alumni", "Apr 2027"),
        ("Coffee Club #5", "As above.", "4 alumni", "Jun 2027"),
        ("Coffee Club #6", "As above — closes the cycle with all cohorts.",
         "4 alumni", "Aug 2027"),
        ("SF · phase 1", "Map the existing circle.", "Circle mapped", "Q1"),
        ("SF · phase 2", "Meet the people who matter in it.", "Key people met",
         "Q1"),
        ("SF · phase 3",
         "Create value there, and stay warm when we need nothing from them. "
         "The phase that gets skipped, and the one that decides whether the "
         "network is real.", "Relationships held", "Q2"),
        ("SF · phase 4", "Extend beyond Avishag's circle. An alum landing in SF "
         "walks into an existing network, not a strange city.",
         "Network beyond her", "Q3–Q4"),
        ("Office Hours · research",
         "How other networks run office hours, and above all how they prevent "
         "abuse, overload and contributor burnout. Check access, matching, "
         "frequency, quality control, fatigue, confidentiality, follow-up. "
         "Bring a recommendation, not a launch.",
         "A recommendation", "Q1, before launch"),
        ("Office Hours · recruit",
         "Quality people willing to give time. Per person: expertise, role, "
         "company, location, what they help with, stage, how often, format, "
         "whether an intro is needed, and their boundaries.",
         "8 by Q1 · 20 by Q2", "Q1–Q2"),
        ("Founder-density dinner",
         "Eight potential founders, not another broad event. §21's own worked "
         "example.", "8 seats", "Q3"),
        ("Month without Avishag",
         "Run the §24 test for real, and answer honestly whether the network "
         "kept strengthening.", "An honest answer", "Q4"),
    ]
    for i, (name, what, target, when) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=name).font = ACCENT
        ws.cell(row=r, column=2, value=what).alignment = WRAP
        ws.cell(row=r, column=3, value=target).alignment = WRAP
        ws.cell(row=r, column=4, value=when)
        for col in (5, 6):
            ws.cell(row=r, column=col).fill = INPUT
        status.add(ws.cell(row=r, column=5))
        for col in range(1, 7):
            c = ws.cell(row=r, column=col)
            c.border = BOX
            if c.font.name != FONT or col != 1:
                c.font = ACCENT if col == 1 else BODY
            if col not in (2, 3, 6):
                c.alignment = TOP
        ws.row_dimensions[r].height = 40

    r = 6 + len(rows)
    ws.cell(row=r, column=1, value="The introduction gate (§3)").font = H2
    ws.cell(row=r + 1, column=1,
            value="Before connecting two people, answer: why do these two need "
                  "to know each other now? No good answer, no introduction. The "
                  "goal is relationships that produce value, not a count of "
                  "intros. And close the loop with the alum — even when Avishag "
                  "does not reply, Delgo is the one who makes sure the circle "
                  "closes.").font = SMALL
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)
    ws.cell(row=r + 1, column=1).alignment = WRAP
    ws.row_dimensions[r + 1].height = 46


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--force", action="store_true",
                    help="overwrite an existing workbook")
    args = ap.parse_args()
    if TARGET.exists() and not args.force:
        sys.exit(f"{TARGET} already exists — anything typed into it would be "
                 f"lost. Pass --force if you mean to rebuild it.")

    roster, schedule, config = load()
    wb = Workbook()
    wb.remove(wb.active)

    n_sched = sheet_schedule(wb, schedule)
    n_roster = sheet_roster(wb, roster)
    sheet_overview(wb, roster, schedule, config, n_roster, n_sched)
    sheet_kpi(wb, config, n_roster, n_sched, schedule)
    sheet_ninety(wb, n_roster, n_sched)
    sheet_projects(wb, config)
    # Reading order: the picture, then today's work, then the people behind it.
    wb._sheets = [wb[name] for name in
                  ["Overview", "Call Schedule", "Roster", "KPI Tracker",
                   "Q1 90-Day Test", "Projects"]]

    TARGET.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TARGET)
    print(f"wrote {TARGET.relative_to(ROOT)}")
    print(f"  {len(roster)} alumni · {len(schedule)} scheduled conversations")
    print(f"  sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
